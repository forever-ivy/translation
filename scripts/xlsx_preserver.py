#!/usr/bin/env python3
"""XLSX format-preserving translation support.

Goal: keep original workbook formatting/structure while replacing translatable text.

Principles:
- Preserve styles, merged cells, column widths, borders, fills, etc.
- Avoid touching numbers/dates/references.
- Optional non-structural beautify: wrap_text + row-height increase only.
"""

from __future__ import annotations

import re
import shutil
from copy import copy as _copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:  # Optional dependency (required for XLSX preservation)
    import openpyxl
except Exception:  # pragma: no cover
    openpyxl = None


_TEXT_HAS_LETTER_RE = re.compile(r"[A-Za-z\u00C0-\u024F\u0370-\u03FF\u0400-\u04FF\u0600-\u06FF\u4E00-\u9FFF]")
_TEXT_ALL_PUNCT_NUM_RE = re.compile(r"^[\s0-9\.,:;()\[\]{}%+\-–—/_\\|]*$")
_TEXT_HAS_ARABIC_RE = re.compile(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]")


@dataclass(frozen=True)
class XlsxCellUnit:
    file: str
    sheet: str
    cell: str
    text: str


def _require_openpyxl() -> None:
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required for XLSX preservation (pip install openpyxl)")


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\u00A0", " ")).strip()


def _is_formula_cell(cell) -> bool:
    try:
        if getattr(cell, "data_type", None) == "f":
            return True
    except Exception:
        pass
    value = getattr(cell, "value", None)
    return isinstance(value, str) and value.strip().startswith("=")


def _is_translatable_text(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    text = _normalize_text(value)
    if not text:
        return False
    if _TEXT_ALL_PUNCT_NUM_RE.fullmatch(text):
        return False
    return bool(_TEXT_HAS_LETTER_RE.search(text))


def extract_translatable_cells(
    xlsx_path: Path,
    *,
    max_cells: int | None = None,
    arabic_only: bool = False,
    interview_only_if_present: bool = False,
    sheet_include_regex: str | None = None,
    sheet_exclude_regex: str | None = None,
    include_formula_display_text: bool = True,
) -> tuple[list[XlsxCellUnit], dict[str, Any]]:
    """Extract translatable cells (string values only) from workbook.

    Returns: (units, meta) where meta includes counts and truncation info.
    """
    _require_openpyxl()
    xlsx_path = Path(xlsx_path).expanduser().resolve()
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    wb_values = None
    if include_formula_display_text:
        # Read cached/display values so formula-backed text cells can be translated.
        wb_values = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    units: list[XlsxCellUnit] = []
    truncated = False

    include_re = None
    exclude_re = None
    if sheet_include_regex and str(sheet_include_regex).strip():
        include_re = re.compile(str(sheet_include_regex))
    if sheet_exclude_regex and str(sheet_exclude_regex).strip():
        exclude_re = re.compile(str(sheet_exclude_regex))

    has_interview = False
    if interview_only_if_present:
        for ws in wb.worksheets:
            if str(ws.title or "").lower().startswith("interview_"):
                has_interview = True
                break

    included_sheets: list[str] = []
    for ws in wb.worksheets:
        title = str(ws.title or "")
        if include_re and not include_re.search(title):
            continue
        if exclude_re and exclude_re.search(title):
            continue
        if has_interview and not title.lower().startswith("interview_"):
            continue
        included_sheets.append(title)

    for ws in wb.worksheets:
        if ws.title not in included_sheets:
            continue
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if _is_formula_cell(cell):
                    if not include_formula_display_text or wb_values is None:
                        continue
                    try:
                        ws_values = wb_values[ws.title]
                        value = ws_values[cell.coordinate].value
                    except Exception:
                        value = None
                if not _is_translatable_text(value):
                    continue
                if arabic_only and not _TEXT_HAS_ARABIC_RE.search(_normalize_text(str(value))):
                    continue
                units.append(
                    XlsxCellUnit(
                        file=xlsx_path.name,
                        sheet=ws.title,
                        cell=str(cell.coordinate),
                        text=_normalize_text(str(value)),
                    )
                )
                if max_cells is not None and len(units) >= max_cells:
                    truncated = True
                    break
            if truncated:
                break
        if truncated:
            break
    wb.close()
    if wb_values is not None:
        wb_values.close()
    return units, {
        "file": xlsx_path.name,
        "cell_count": len(units),
        "truncated": truncated,
        "max_cells": max_cells,
        "arabic_only": bool(arabic_only),
        "interview_only_if_present": bool(interview_only_if_present),
        "sheet_include_regex": str(sheet_include_regex or ""),
        "sheet_exclude_regex": str(sheet_exclude_regex or ""),
        "include_formula_display_text": bool(include_formula_display_text),
        "included_sheets": included_sheets,
    }


def units_to_payload(units: Iterable[XlsxCellUnit], *, max_chars_per_cell: int = 400) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for unit in units:
        text = unit.text
        if max_chars_per_cell > 0 and len(text) > max_chars_per_cell:
            text = text[:max_chars_per_cell]
        out.append(
            {
                "file": unit.file,
                "sheet": unit.sheet,
                "cell": unit.cell,
                "text": text,
            }
        )
    return out


def _normalize_xlsx_translation_map(entries: Any, *, file_hint: str | None = None) -> dict[tuple[str, str], str]:
    """Normalize model output into {(sheet, cell): text} for a specific file."""
    out: dict[tuple[str, str], str] = {}
    if not entries:
        return out
    if isinstance(entries, dict):
        # Accept {"Sheet1!B2": "text"} shape as a convenience
        for k, v in entries.items():
            key = str(k or "")
            if "!" in key:
                sheet, cell = key.split("!", 1)
                out[(sheet.strip(), cell.strip().upper())] = str(v or "")
        return out
    if not isinstance(entries, list):
        return out

    for item in entries:
        if not isinstance(item, dict):
            continue
        file_value = str(item.get("file") or "").strip()
        if file_hint and file_value and file_value != file_hint:
            continue
        sheet = str(item.get("sheet") or "").strip()
        cell = str(item.get("cell") or "").strip().upper()
        text = str(item.get("text") or "")
        if not sheet or not cell:
            continue
        out[(sheet, cell)] = text
    return out


def beautify_xlsx_non_structural(
    xlsx_path: Path,
    *,
    changed_cells: Iterable[tuple[str, str]],
    base_row_height: float = 15.0,
) -> None:
    """Non-structural beautify: wrap_text + row height adjustments only."""
    _require_openpyxl()
    xlsx_path = Path(xlsx_path).expanduser().resolve()
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=False)
    touched_rows: set[tuple[str, int]] = set()

    for sheet, cell_addr in changed_cells:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        try:
            cell = ws[cell_addr]
        except Exception:
            continue

        align = cell.alignment
        try:
            new_align = _copy(align)
            new_align.wrap_text = True
            cell.alignment = new_align
        except Exception:
            # Fallback: set wrap_text only when copy not available
            cell.alignment = type(align)(**{**align.__dict__, "wrap_text": True})  # pragma: no cover

        touched_rows.add((sheet, int(cell.row)))

    # Increase row height for touched rows only (heuristic).
    for sheet, row_idx in touched_rows:
        ws = wb[sheet]
        dim = ws.row_dimensions[row_idx]
        current = float(dim.height) if dim.height is not None else base_row_height
        # Conservative bump to reduce clipping when translations are longer.
        dim.height = max(current, base_row_height * 1.35)

    wb.save(str(xlsx_path))
    wb.close()


def apply_translation_map(
    *,
    source_xlsx: Path,
    output_xlsx: Path,
    translation_map_entries: Any,
    beautify: bool = True,
    overwrite_formula_cells: bool = True,
) -> dict[str, Any]:
    """Copy source workbook and apply translations by (sheet, cell)."""
    _require_openpyxl()
    source_xlsx = Path(source_xlsx).expanduser().resolve()
    output_xlsx = Path(output_xlsx).expanduser().resolve()
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(source_xlsx), str(output_xlsx))

    mapped = _normalize_xlsx_translation_map(translation_map_entries, file_hint=source_xlsx.name)
    wb = openpyxl.load_workbook(str(output_xlsx), data_only=False)

    changed: list[tuple[str, str]] = []
    missing_sheets = 0
    missing_cells = 0
    skipped_formulas = 0

    for (sheet, cell_addr), new_text in mapped.items():
        if sheet not in wb.sheetnames:
            missing_sheets += 1
            continue
        ws = wb[sheet]
        try:
            cell = ws[cell_addr]
        except Exception:
            missing_cells += 1
            continue
        if _is_formula_cell(cell):
            if not overwrite_formula_cells:
                skipped_formulas += 1
                continue
            # Explicitly mapped formula-backed text cell: replace formula with translated static text.
            cell.value = str(new_text)
            changed.append((sheet, cell_addr))
            continue
        if not _is_translatable_text(cell.value) and cell.value is not None:
            # Avoid turning non-text cells into strings.
            continue
        cell.value = str(new_text)
        changed.append((sheet, cell_addr))

    wb.save(str(output_xlsx))
    wb.close()

    if beautify and changed:
        beautify_xlsx_non_structural(output_xlsx, changed_cells=changed)

    return {
        "ok": True,
        "source": str(source_xlsx),
        "output": str(output_xlsx),
        "applied_count": len(changed),
        "missing_sheets": missing_sheets,
        "missing_cells": missing_cells,
        "skipped_formulas": skipped_formulas,
    }
