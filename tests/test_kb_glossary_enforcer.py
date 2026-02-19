#!/usr/bin/env python3

import tempfile
import unittest
from pathlib import Path

from docx import Document
from openpyxl import Workbook

from scripts.kb_glossary_enforcer import (
    build_glossary_map,
    load_company_glossary_pairs,
    select_terms_for_sources,
)


def _make_xlsx(path: Path, *, rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in rows:
        ws.append(r)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _make_docx_table(path: Path, *, rows: list[tuple[str, str]]) -> None:
    doc = Document()
    table = doc.add_table(rows=len(rows), cols=2)
    for idx, (ar, en) in enumerate(rows):
        table.cell(idx, 0).text = ar
        table.cell(idx, 1).text = en
    path.parent.mkdir(parents=True, exist_ok=True)
    doc.save(path)


class KbGlossaryEnforcerTest(unittest.TestCase):
    def test_load_pairs_from_company_glossary_xlsx_and_docx(self):
        with tempfile.TemporaryDirectory() as tmp:
            kb_root = Path(tmp) / "Knowledge Repository"
            company_dir = kb_root / "00_Glossary" / "Eventranz"

            _make_xlsx(
                company_dir / "glossary.xlsx",
                rows=[
                    ["العربية", "English"],  # header row should be skipped
                    ["مدرسة", "School"],
                ],
            )
            _make_docx_table(company_dir / "glossary.docx", rows=[("جامعة", "University")])

            pairs, meta = load_company_glossary_pairs(kb_root=kb_root, company="Eventranz")
            self.assertFalse(meta.get("skipped_missing_dir"))
            self.assertGreaterEqual(meta.get("files_scanned", 0), 2)
            self.assertGreaterEqual(len(pairs), 2)

            glossary_map, conflicts = build_glossary_map(pairs)
            self.assertIn("مدرسة", [p.arabic for p in glossary_map.values()])
            self.assertIn("جامعة", [p.arabic for p in glossary_map.values()])
            self.assertEqual(conflicts, [])

    def test_conflict_resolution_prefers_longer_english(self):
        with tempfile.TemporaryDirectory() as tmp:
            kb_root = Path(tmp) / "Knowledge Repository"
            company_dir = kb_root / "00_Glossary" / "Eventranz"

            _make_xlsx(
                company_dir / "g.xlsx",
                rows=[
                    ["مدرسة", "School"],
                    ["مدرسة", "Primary School"],  # longer should win
                ],
            )

            pairs, _meta = load_company_glossary_pairs(kb_root=kb_root, company="Eventranz")
            glossary_map, conflicts = build_glossary_map(pairs)
            self.assertTrue(conflicts)
            self.assertEqual(glossary_map[next(iter(glossary_map.keys()))].english, "Primary School")

    def test_select_terms_matches_only_present_terms(self):
        with tempfile.TemporaryDirectory() as tmp:
            kb_root = Path(tmp) / "Knowledge Repository"
            company_dir = kb_root / "00_Glossary" / "Eventranz"

            _make_xlsx(company_dir / "g.xlsx", rows=[["مدرسة", "School"], ["جامعة", "University"]])
            pairs, _meta = load_company_glossary_pairs(kb_root=kb_root, company="Eventranz")
            glossary_map, _conflicts = build_glossary_map(pairs)

            selected, sel_meta = select_terms_for_sources(
                glossary_map=glossary_map,
                source_texts=["هذا مدرسة جميلة"],
                max_terms=80,
            )
            self.assertEqual(sel_meta.get("matched_terms"), 1)
            self.assertEqual(selected[0].arabic, "مدرسة")
            self.assertEqual(selected[0].english, "School")


if __name__ == "__main__":
    unittest.main()

