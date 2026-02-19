#!/usr/bin/env python3

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from scripts.xlsx_preserver import apply_translation_map
from scripts.xlsx_preserver import extract_translatable_cells


def _make_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Hello"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="00FFDD88")
    ws.column_dimensions["A"].width = 20

    ws["A2"] = "=SUM(1,2)"
    ws["B1"] = 123

    ws.merge_cells("C1:D1")
    ws["C1"] = "Merged"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_interview_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Interview_Schools"
    ws["A1"] = "مرحبا"
    ws["A2"] = "Hello"
    ws["A3"] = "=SUM(1,2)"

    lookups = wb.create_sheet(title="Lookups")
    lookups["A1"] = "مرحبا"
    lookups["A2"] = "Lookup"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


def _make_non_interview_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "مرحبا"
    ws2 = wb.create_sheet(title="Lookups")
    ws2["A1"] = "مرحبا"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    wb.close()


class XlsxPreserverTest(unittest.TestCase):
    def test_apply_translation_preserves_styles_merges_and_skips_formulas(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            out = Path(tmp) / "out.xlsx"
            _make_xlsx(src)

            res = apply_translation_map(
                source_xlsx=src,
                output_xlsx=out,
                translation_map_entries=[
                    {"file": src.name, "sheet": "Sheet1", "cell": "A1", "text": "Bonjour"},
                    {"file": src.name, "sheet": "Sheet1", "cell": "A2", "text": "SHOULD_NOT_APPLY"},
                    {"file": src.name, "sheet": "Sheet1", "cell": "B1", "text": "999"},
                ],
                beautify=False,
                overwrite_formula_cells=False,
            )
            self.assertTrue(res.get("ok"))
            self.assertEqual(res.get("applied_count"), 1)
            self.assertEqual(res.get("skipped_formulas"), 1)

            wb = load_workbook(str(out), data_only=False)
            ws = wb["Sheet1"]

            self.assertEqual(ws["A1"].value, "Bonjour")
            self.assertTrue(ws["A1"].font.bold)
            self.assertEqual(ws.column_dimensions["A"].width, 20)
            self.assertEqual(ws["A2"].value, "=SUM(1,2)")
            self.assertEqual(ws["B1"].value, 123)
            self.assertTrue(any(str(rng) == "C1:D1" for rng in ws.merged_cells.ranges))

            wb.close()

    def test_beautify_enables_wrap_text_and_bumps_row_height(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            out = Path(tmp) / "out.xlsx"
            _make_xlsx(src)

            apply_translation_map(
                source_xlsx=src,
                output_xlsx=out,
                translation_map_entries=[{"file": src.name, "sheet": "Sheet1", "cell": "A1", "text": "Bonjour le monde"}],
                beautify=True,
            )

            wb = load_workbook(str(out), data_only=False)
            ws = wb["Sheet1"]
            self.assertTrue(ws["A1"].alignment.wrap_text)
            self.assertIsNotNone(ws.row_dimensions[1].height)
            self.assertGreaterEqual(float(ws.row_dimensions[1].height), 20.0)
            wb.close()

    def test_apply_translation_can_overwrite_formula_cells_when_enabled(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            out = Path(tmp) / "out.xlsx"
            _make_xlsx(src)

            res = apply_translation_map(
                source_xlsx=src,
                output_xlsx=out,
                translation_map_entries=[{"file": src.name, "sheet": "Sheet1", "cell": "A2", "text": "Three"}],
                beautify=False,
                overwrite_formula_cells=True,
            )
            self.assertTrue(res.get("ok"))
            self.assertEqual(res.get("applied_count"), 1)
            self.assertEqual(res.get("skipped_formulas"), 0)

            wb = load_workbook(str(out), data_only=False)
            ws = wb["Sheet1"]
            self.assertEqual(ws["A2"].value, "Three")
            wb.close()

    def test_extract_translatable_cells_arabic_only_interview_focus(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            _make_interview_workbook(src)

            units, meta = extract_translatable_cells(
                src,
                arabic_only=True,
                interview_only_if_present=True,
            )
            self.assertEqual([u.sheet for u in units], ["Interview_Schools"])
            self.assertEqual([u.cell for u in units], ["A1"])
            self.assertEqual([u.text for u in units], ["مرحبا"])
            self.assertTrue(meta.get("arabic_only"))
            self.assertTrue(meta.get("interview_only_if_present"))
            self.assertIn("Interview_Schools", meta.get("included_sheets") or [])
            self.assertNotIn("Lookups", meta.get("included_sheets") or [])

    def test_extract_translatable_cells_interview_focus_falls_back_when_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            _make_non_interview_workbook(src)

            units, meta = extract_translatable_cells(
                src,
                arabic_only=True,
                interview_only_if_present=True,
            )
            self.assertEqual(len(units), 2)
            self.assertCountEqual([u.sheet for u in units], ["Sheet1", "Lookups"])
            self.assertIn("Sheet1", meta.get("included_sheets") or [])
            self.assertIn("Lookups", meta.get("included_sheets") or [])


if __name__ == "__main__":
    unittest.main()
